from openpyxl import Workbook , load_workbook
"""
try:
    wd = load_workbook("C:/Users/siddh/OneDrive/Desktop/python ai ml/program/excel x python/sales data.xlsx")
    print(wd.sheetnames)
    ws= wd.active
    Total_cost = 0
    for row in range(2 , ws.max_row +1 ):
        cost = ws[f"C{row}"].value
        #print(cost)
        if isinstance(cost , (int , float)):
            Total_cost = Total_cost + cost
            print(Total_cost)
except Exception as e :
    print(e)
"""
"""
try :
    wd = load_workbook("C:/Users/siddh/OneDrive/Desktop/python ai ml/program/excel x python/sales data.xlsx")
    ws = wd.active 
    ws["K1"] = "tax"
    for row in range(2 , ws.max_row+1):
        cost = ws[f'C{row}'].value
        if isinstance(cost , (int , float)):
            tax = cost * 0.18
            ws['K{}'.format(row)] = tax
    wd.save('C:/Users/siddh/OneDrive/Desktop/python ai ml/program/excel x python/updated sales data.xlsx')
except Exception as e:
    print(e)
"""

