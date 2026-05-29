from openpyxl import Workbook , load_workbook
"""
try:
    wb = load_workbook("C:/Users/siddh/OneDrive/Desktop/python ai ml/program/excel x python/Inventory-Records-Sample-Data.xlsx")
    ws = wb.active
    total_scale = 0
    print(ws['D7'].value)
    print(ws.values)
    for row in ws.iter_rows(min_row=7, values_only=True):
        total_scale = total_scale + row[8]
    print(total_scale)
    for row in ws.iter_rows(min_row=7,values_only=True):
        total_scale = total_scale + row[7]
    print(total_scale)
  #  ws['J6'] = "tax"
    for row in range(7, ws.max_row+1):
        costx = ws[f"I{row}"].value
        #print(costx)
        if isinstance(costx ,(int , float)):
            tax = costx *0.18 
            ws[f"J{row}"] = tax
    wb.save("updated.xlsx")
except Exception as e:
    print(e)
    
"""

try :
    wb = load_workbook("C:/Users/siddh/OneDrive/Desktop/python ai ml/program/excel x python/Inventory-Records-Sample-Data.xlsx")
    ws = wb.active
    ws['K6'] = "platform cost"
    wb.save('updated2.xlsx')
    for row in range(7 , ws.max_row+1):
        cost = ws[f"I{row}"].value
        if isinstance(cost , (int , float)):
            e_cost = cost + 50 
            ws[f"K{row}"] = e_cost
    ws.delete_cols(9)
    wb.save('C:/Users/siddh/OneDrive/Desktop/python ai ml/program/excel x python/updated.xlsx')
except Exception as e:
    print(e)

    