from openpyxl import Workbook , load_workbook
"""
data = [
  {
    "id": 1,
    "name": "Anshu",
    "age": 19,
    "city": "Bhopal",
    "course": "Python ML",
    "marks": 88
  },
  {
    "id": 2,
    "name": "Rahul",
    "age": 21,
    "city": "Indore",
    "course": "Data Science",
    "marks": 72
  },
  {
    "id": 3,
    "name": "Priya",
    "age": 20,
    "city": "Delhi",
    "course": "AI",
    "marks": 45
  },
  {
    "id": 4,
    "name": "Karan",
    "age": 22,
    "city": "Mumbai",
    "course": "Web Dev",
    "marks": 91
  },
  {
    "id": 5,
    "name": "Sneha",
    "age": 18,
    "city": "Pune",
    "course": "Cyber Security",
    "marks": 67
  }
]

try:
    wd = Workbook()
    ws = wd.active
    ws.title = "sec"

    ws.append(["id","name","age","city","course","marks"])
    for x in data :
        row = [
            x["id"],
            x["name"],
            x["age"],
            x["city"],
            x["course"],
            x["marks"]
        ]
        ws.append(row)

    wd.save("test2.xlsx")
except Exception as e:
    print(e)

print(ws['B2'].value)
print(ws['A1'])
ws['B3'] = "invoice"


print(wd.sheetnames)

try :
  wd = load_workbook("C:/Users/siddh/OneDrive/Desktop/python ai ml/program/test1.xlsx")
  ws = wd.active
  
  ws = wd['sheet1']
  print(ws['B2'].value)
  ws['B3'] = 'yash'
  wd.save('test1.xlsx')
except Exception as e:
  print(e)
"""

try :
  wd = load_workbook("C:/Users/siddh/OneDrive/Desktop/python ai ml/program/test1.xlsx")
  ws = wd['sheet1']
  for i in ws.iter_rows(values_only=True):
    print(i[1],i[2],i[3],i[4])
  for j in ws.iter_cols(values_only=True):
    print(j)
except Exception as e:
  print(e)
